from django.shortcuts import render, redirect, get_object_or_404
from datetime import datetime, timedelta
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from .models import Sale, SalaryPayment, CrmUser, Company, ImportLock, Employee, ExpenseType, ProductionExpense
from django.contrib import messages
from django.db.models import Sum, IntegerField, Value, CharField, F
from django.http import JsonResponse, HttpResponse
from django.db.models.functions import Cast, Concat, TruncMonth
from django.views.decorators.http import require_GET
from collections import defaultdict
import requests
from requests.exceptions import RequestException, Timeout
from decimal import Decimal, InvalidOperation
from dateutil.parser import parse as parse_date
import openpyxl
from django.db import transaction
from django.utils.timezone import now
from django.utils.translation import gettext_lazy
from .forms import RegistrationForm, SalaryPaymentForm, ProductionExpenseForm, EmployeeForm, ExpenseTypeForm
from .decorators import admin_required
from .utils import (
    get_months,
    get_current_crm_user,
    normalize_amount,
    get_month_date_range,
    apply_dashboard_filters,
    parse_date_range,
)
# from . import ai_views  # DISABLED: local AI features are turned off
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from urllib.parse import quote
from django.conf import settings
from django.urls import reverse
import logging
from urllib.parse import urlparse
from typing import cast, Any
from django.utils.encoding import force_str
import re

logger = logging.getLogger(__name__)

def is_ajax(request):
    return request.headers.get('x-requested-with') == 'XMLHttpRequest'



def _webhook_urls():
    base = settings.CRM_WEBHOOK_BASE
    return {
        'deals': f'{base}/crm.deal.list',
        'users': f'{base}/user.get',
        'companies': f'{base}/crm.company.list',
    }


def _post_with_retry(url, *, json, timeout_seconds=None, connect_timeout_seconds=None, max_retries=3, on_retry=None):
    """
    Timeout (connect, read): for slow VPN/SSL or heavy CRM we allow more time.
    By default values are taken from settings (CRM_CONNECT_TIMEOUT, CRM_READ_TIMEOUT).
    """
    if connect_timeout_seconds is None:
        connect_timeout_seconds = getattr(settings, 'CRM_CONNECT_TIMEOUT', 30)
    if timeout_seconds is None:
        timeout_seconds = getattr(settings, 'CRM_READ_TIMEOUT', 120)
    last_exc = None
    timeout_tuple = (connect_timeout_seconds, timeout_seconds)
    for attempt in range(1, max_retries + 1):
        try:
            resp = requests.post(url, json=json, timeout=timeout_tuple)
            resp.raise_for_status()
            return resp
        except (Timeout, RequestException, requests.exceptions.ConnectionError) as exc:
            last_exc = exc
            if callable(on_retry):
                try:
                    on_retry(attempt, max_retries)
                except Exception:
                    pass
            logger.warning("CRM POST failed (attempt %s/%s): %s", attempt, max_retries, exc)
            if attempt == max_retries:
                break
    raise last_exc if last_exc is not None else RequestException("Max retries exceeded")


def _validate_webhook_base():
    base = getattr(settings, 'CRM_WEBHOOK_BASE', '') or ''
    try:
        parsed = urlparse(base)
    except Exception:
        return "CRM_WEBHOOK_BASE: invalid URL. Please check your .env configuration."

    if not parsed.scheme or not parsed.netloc:
        return "CRM_WEBHOOK_BASE: missing scheme or domain. Example: https://your.crm24.ru/rest/1/token"

    host = parsed.hostname or ''
    path = parsed.path or ''

    if 'example.local' in host or 'placeholder' in base:
        return "Please specify a real CRM24 webhook in .env (CRM_WEBHOOK_BASE). A placeholder is currently set."

    if '/rest/' not in path:
        return "CRM_WEBHOOK_BASE must contain a path of the form /rest/<userId>/<token>."

    return None


@login_required
@require_GET
def import_status(request):
    """Import status: sales import progress in % and remaining time (for frontend polling)."""
    lock, _ = ImportLock.objects.get_or_create(id=1)
    # If the lock is "stuck" (import interrupted, server restarted) — assume there is no active import
    stale_minutes = 2
    if lock.is_locked:
        if not lock.updated_at or (timezone.now() - lock.updated_at > timedelta(minutes=stale_minutes)):
            ImportLock.objects.filter(id=1).update(is_locked=False, updated_at=timezone.now())
            lock = ImportLock.objects.get(id=1)
    is_locked = bool(lock.is_locked)
    return JsonResponse({
        "is_locked": is_locked,
        "stage": getattr(lock, "stage", "idle"),
        "message": getattr(lock, "message", ""),
        "progress_percent": getattr(lock, "progress_percent", 0.0),
        "processed": getattr(lock, "processed", 0),
        "total": getattr(lock, "total", 0),
        "eta_seconds": getattr(lock, "eta_seconds", 0),
        "started_at": lock.started_at.isoformat() if getattr(lock, "started_at", None) else None,
    })


@login_required
def index(request):
    accept_header = request.headers.get("Accept", "")
    wants_json = "application/json" in accept_header

    if request.method != "POST":
        if wants_json:
            return JsonResponse({"success": False, "error": gettext_lazy("Method not allowed")}, status=405)
        return render(request, "salary/error_simple.html", {
            "message": gettext_lazy("Method not allowed. Return to the sales page and click 'Load'."),
        }, status=405)

    settings_error = _validate_webhook_base()
    if settings_error:
        if wants_json:
            return JsonResponse({"success": False, "error": settings_error})
        return render(request, "salary/error_simple.html", {"message": settings_error}, status=200)

    with transaction.atomic():
        lock, _ = ImportLock.objects.select_for_update().get_or_create(id=1)
        if lock.is_locked and (now() - lock.updated_at < timedelta(minutes=5)):
            msg = gettext_lazy("Import is already running by another user. Please try again later.")
            if wants_json:
                return JsonResponse({"success": False, "error": msg})
            return render(request, "salary/error_simple.html", {"message": msg}, status=200)
        lock.is_locked = True
        lock.updated_at = now()
        lock.stage = "starting"
        lock.message = gettext_lazy("Preparing…")
        lock.progress_percent = 0.0
        lock.processed = 0
        lock.total = 0
        lock.eta_seconds = 0
        lock.started_at = timezone.now()
        lock.save()

    try:
        ImportLock.objects.filter(id=1).update(
            stage="users", message=gettext_lazy("Loading users…"),
            progress_percent=5.0, updated_at=timezone.now()
        )
        user_map = import_users()
        ImportLock.objects.filter(id=1).update(
            stage="companies", message=gettext_lazy("Loading companies…"),
            progress_percent=10.0, updated_at=timezone.now()
        )
        company_map = import_companies()
        sales_to_create = []
        start = 0
        existing_ids = set(Sale.objects.values_list('id_number', flat=True))
        total_deals = 0
        started_at = timezone.now()

        ImportLock.objects.filter(id=1).update(
            stage="deals", message=gettext_lazy("Loading deals…"),
            progress_percent=15.0, updated_at=timezone.now()
        )

        deal_batch_num = 0
        while True:
            ImportLock.objects.filter(id=1).update(
                stage="deals", message=gettext_lazy("Loading deals… (page %s)") % (deal_batch_num + 1),
                progress_percent=min(99.0, 15.0 + deal_batch_num * 0.3), updated_at=timezone.now()
            )
            data = {
                'start': start,
                'select': [
                    'ID', 'ASSIGNED_BY_ID', 'STAGE_ID', 'OPPORTUNITY',
                    'CLOSEDATE', 'COMPANY_ID', 'UF_CRM_1736157032', 'UF_CRM_1740138171',
                    'TITLE'
                ]
            }
            def _on_retry_deals(attempt, _max):
                ImportLock.objects.filter(id=1).update(
                    stage="deals", message=gettext_lazy("Loading deals… (retry %s/%s)") % (attempt, _max),
                    progress_percent=min(99.0, 15.0 + deal_batch_num * 0.3 + attempt * 0.2), updated_at=timezone.now()
                )
            response = _post_with_retry(_webhook_urls()['deals'], json=data, on_retry=_on_retry_deals)
            response_json = response.json()
            if total_deals == 0 and response_json.get("total") is not None:
                try:
                    total_deals = int(response_json["total"])
                except (TypeError, ValueError):
                    pass
            batch = response_json.get('result', [])
            if not batch:
                break

            processed = start + len(batch)
            if "next" in response_json:
                try:
                    processed = int(response_json["next"])
                except (TypeError, ValueError):
                    pass
            percent = 15.0
            eta_seconds = 0
            if total_deals > 0:
                percent = min(99.0, 15.0 + (processed / total_deals) * 80.0)
                elapsed = (timezone.now() - started_at).total_seconds()
                if elapsed > 0 and processed > 0:
                    rate = processed / elapsed
                    if rate > 0:
                        eta_seconds = int((total_deals - processed) / rate)
            ImportLock.objects.filter(id=1).update(
                stage="deals", message=gettext_lazy("Loading deals…"),
                progress_percent=percent, processed=processed, total=total_deals,
                eta_seconds=eta_seconds, updated_at=timezone.now()
            )
            deal_batch_num += 1

            for item in batch:
                if item.get('STAGE_ID') != "C1:WON":
                    continue
                deal_id = str(item.get('ID'))
                if deal_id in existing_ids:
                    continue
                manager = user_map.get(int(item.get('ASSIGNED_BY_ID')))
                if not manager:
                    continue
                try:
                    closing_date = parse_date(item.get("CLOSEDATE")).date() if item.get("CLOSEDATE") else None
                except (ValueError, TypeError):
                    closing_date = None
                company_id = item.get('COMPANY_ID')
                company = company_map.get(int(company_id)) if company_id else None
                sales_to_create.append(Sale(
                    id_number=deal_id,
                    manager=manager,
                    sale=item.get('OPPORTUNITY') or 0,
                    company=company,
                    account_number=(
                        item.get('UF_CRM_1740138171')[0]
                        if isinstance(item.get('UF_CRM_1740138171'), list) and item.get('UF_CRM_1740138171')
                        else item.get('UF_CRM_1740138171') or ""
                    ),
                    salary=parse_decimal(item.get('UF_CRM_1736157032')),
                    closing_date=closing_date,
                    title=item.get('TITLE') or '',
                ))

            if 'next' in response_json:
                start = response_json['next']
            else:
                break

        if sales_to_create:
            Sale.objects.bulk_create(sales_to_create, ignore_conflicts=True)
        ImportLock.objects.filter(id=1).update(
            stage="done", message=gettext_lazy("Done"),
            progress_percent=100.0, eta_seconds=0, updated_at=timezone.now()
        )
        success_message = gettext_lazy("Update completed. %(count)s new deals have been added.") % {
            "count": len(sales_to_create)
        }
        if wants_json:
            return JsonResponse({"success": True, "message": success_message})
        return render(request, "salary/error_simple.html", {"message": success_message}, status=200)

    except (RequestException, Timeout) as e:
        logger.warning("Import failed (CRM unreachable): %s", e)
        error_message = gettext_lazy("Failed to update sales. Please check integration settings and try again.")
        ImportLock.objects.filter(id=1).update(
            stage="error", message=error_message,
            eta_seconds=0, updated_at=timezone.now()
        )
        if wants_json:
            return JsonResponse({"success": False, "error": error_message}, status=200)
        return render(request, "salary/error_simple.html", {"message": error_message}, status=200)
    except Exception as e:
        logger.exception("Import failed: %s", e)
        error_message = gettext_lazy("Failed to update sales. Please check integration settings and try again.")
        ImportLock.objects.filter(id=1).update(
            stage="error", message=error_message,
            eta_seconds=0, updated_at=timezone.now()
        )
        if wants_json:
            return JsonResponse({"success": False, "error": error_message}, status=200)
        return render(request, "salary/error_simple.html", {"message": error_message}, status=200)
    finally:
        ImportLock.objects.filter(id=1).update(is_locked=False, updated_at=timezone.now())


def import_users():
    """
    Imports or updates users from CRM24.
    PERFORMANCE: Uses bulk operations to minimize database queries.
    Returns a dictionary mapping user_id to CrmUser object for quick lookups.
    """
    try:
        start = 0
        existing_users = {u.user_id: u for u in CrmUser.objects.all()}
        users_to_create = []
        users_to_update = []
        batch_num = 0

        while True:
            ImportLock.objects.filter(id=1).update(
                stage="users", message="Loading users… (page %s)" % (batch_num + 1),
                progress_percent=min(9.0, 5.0 + batch_num * 0.4), updated_at=timezone.now()
            )
            payload = {"start": start}
            def _on_retry_users(attempt, _max):
                ImportLock.objects.filter(id=1).update(
                    stage="users", message="Loading users… (retry %s/%s)" % (attempt, _max),
                    progress_percent=min(9.0, 5.0 + batch_num * 0.4 + attempt * 0.15), updated_at=timezone.now()
                )
            response = _post_with_retry(_webhook_urls()['users'], json=payload, on_retry=_on_retry_users)
            api_users = response.json().get("result", [])
            if not api_users:
                break

            for u in api_users:
                user_id = int(u["ID"])
                user_data = {
                    "name": u.get("NAME") or "",
                    "last_name": u.get("LAST_NAME") or "",
                }
                if user_id in existing_users:
                    existing_user = existing_users[user_id]
                    if existing_user.name != user_data['name'] or existing_user.last_name != user_data['last_name']:
                        existing_user.name = user_data['name']
                        existing_user.last_name = user_data['last_name']
                        users_to_update.append(existing_user)
                else:
                    users_to_create.append(CrmUser(user_id=user_id, **user_data))

            batch_num += 1
            progress = min(9.0, 5.0 + batch_num * 1.5)
            ImportLock.objects.filter(id=1).update(
                stage="users", message="Loading users…",
                progress_percent=progress, updated_at=timezone.now()
            )

            if "next" in response.json():
                start = response.json()["next"]
            else:
                break

        if users_to_create:
            CrmUser.objects.bulk_create(users_to_create)
        if users_to_update:
            CrmUser.objects.bulk_update(users_to_update, ['name', 'last_name'])

        logger.info("User import completed")
        # Return a fresh map of all users
        return {u.user_id: u for u in CrmUser.objects.all()}
    except (RequestException, Timeout) as e:
        logger.warning("Error during user import (network/timeout): %s", e)
        raise
    except Exception as e:
        logger.exception("Error during user import: %s", e)
        raise


def import_companies():
    """
    Imports or updates companies from CRM24.
    PERFORMANCE: Uses bulk operations to minimize database queries.
    Returns a dictionary mapping company_id to Company object for quick lookups.
    """
    try:
        start = 0
        existing_companies = {c.company_id: c for c in Company.objects.all()}
        companies_to_create = []
        companies_to_update = []
        batch_num = 0

        while True:
            ImportLock.objects.filter(id=1).update(
                stage="companies", message="Loading companies… (page %s)" % (batch_num + 1),
                progress_percent=min(14.0, 10.0 + batch_num * 0.4), updated_at=timezone.now()
            )
            payload = {"start": start, "select": ["ID", "TITLE"]}
            def _on_retry_companies(attempt, _max):
                ImportLock.objects.filter(id=1).update(
                    stage="companies", message="Loading companies… (retry %s/%s)" % (attempt, _max),
                    progress_percent=min(14.0, 10.0 + batch_num * 0.4 + attempt * 0.2), updated_at=timezone.now()
                )
            response = _post_with_retry(_webhook_urls()['companies'], json=payload, on_retry=_on_retry_companies)
            api_companies = response.json().get("result", [])
            if not api_companies:
                break

            for c in api_companies:
                company_id = int(c["ID"])
                title = c.get("TITLE") or ""
                if company_id in existing_companies:
                    if existing_companies[company_id].title != title:
                        existing_companies[company_id].title = title
                        companies_to_update.append(existing_companies[company_id])
                else:
                    companies_to_create.append(Company(company_id=company_id, title=title))

            batch_num += 1
            progress = min(14.0, 10.0 + batch_num * 1.5)
            ImportLock.objects.filter(id=1).update(
                stage="companies", message="Loading companies…",
                progress_percent=progress, updated_at=timezone.now()
            )

            if "next" in response.json():
                start = response.json()["next"]
            else:
                break

        if companies_to_create:
            Company.objects.bulk_create(companies_to_create)
        if companies_to_update:
            Company.objects.bulk_update(companies_to_update, ['title'])

        logger.info("Company import completed")
        return {c.company_id: c for c in Company.objects.all()}
    except (RequestException, Timeout) as e:
        logger.warning("Error during company import (network/timeout): %s", e)
        raise
    except Exception as e:
        logger.exception("Error during company import: %s", e)
        raise

def parse_decimal(value):
    try:
        if isinstance(value, list):
            value = value[0] if value else ''
        value = str(value).replace(" ", "").replace("М", "").replace(",", ".").strip()
        return Decimal(value)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal("0.0")
    

@login_required
def sales_list(request):
    # Filter parameters
    month = request.GET.get('month')
    year_param = request.GET.get('year')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    filter_type = request.GET.get('filter_type', 'month')
    sort_by = request.GET.get('sort', 'closing_date')
    order = request.GET.get('order', 'desc')

    months = get_months()
    current_user, is_admin = get_current_crm_user(request)

    # Base querysets
    sales = Sale.objects.select_related('manager', 'company').annotate(
        id_as_int=Cast('id_number', IntegerField())
    )

    # Whitelist and mapping for sorting
    sortable_fields = {
        'id': 'id_as_int',
        'manager': 'manager__last_name',
        'company': 'company__title',
        'account_number': 'account_number',
        'closing_date': 'closing_date',
        'sale': 'sale',
        'salary': 'salary',
    }

    # Default sort
    final_order_by = ['-closing_date', '-id_as_int']

    if sort_by in sortable_fields:
        sort_field = sortable_fields[sort_by]
        prefix = '-' if order == 'desc' else ''
        final_order_by = [f'{prefix}{sort_field}', '-id_as_int']

    sales = sales.order_by(*final_order_by)


    salary_qs = SalaryPayment.objects.all()
    manager_id = request.GET.get('manager')
    expenses_qs = ProductionExpense.objects.all()
    sales, salary_qs, expenses_qs, _dash_meta = apply_dashboard_filters(
        sales,
        salary_qs,
        expenses_qs,
        is_admin=is_admin,
        current_user=current_user,
        manager_id=manager_id,
        month=month,
        year_param=year_param,
        date_from=date_from,
        date_to=date_to,
        filter_type=filter_type,
    )
    selected_year = _dash_meta.get("effective_year") or ""

    # -------------------------------
    # Aggregations and groupings
    # -------------------------------

    total_salary = sales.aggregate(total=Sum('salary'))['total'] or 0
    total_sales = sales.aggregate(total=Sum('sale'))['total'] or 0

    total_expenses = expenses_qs.aggregate(total=Sum('amount'))['total'] or 0

    # The salary_qs is now correctly filtered by the same period as sales
    salary_paid = salary_qs.aggregate(total=Sum('amount'))['total'] or 0
    salary_left = total_salary - salary_paid

    grouped_sales = defaultdict(lambda: {'sales': [], 'total_sale': 0, 'total_salary': 0})
    # The sales queryset is already filtered, so we just group it for display.
    for sale in sales:
        if sale.closing_date:
            ym_key = (sale.closing_date.year, sale.closing_date.month)
            row = cast(dict[str, Any], grouped_sales[ym_key])
            row['sales'].append(sale)
            row['total_sale'] += sale.sale or 0
            row['total_salary'] += sale.salary or 0

    grouped_sales = dict(sorted(grouped_sales.items()))

    # --- Chart Data Preparation ---

    # Helper function and dict for sorting month labels
    month_name_to_num = {v: k for k, v in months.items()}
    def month_year_key(m):
        try:
            month_str, year_str = m.split()
            month_num = month_name_to_num.get(month_str, 0)
            year_num = int(year_str)
            return (year_num, month_num)
        except Exception:
            return (0, 0)

    # --- Chart 1: Sales & Salary by Month (using DB aggregation) ---
    chart_agg = sales.annotate(
        month_group=TruncMonth('closing_date')
    ).values('month_group').annotate(
        total_sale=Sum('sale'),
        total_salary=Sum('salary')
    ).order_by('month_group')

    chart_labels = [f"{months[agg['month_group'].month]} {agg['month_group'].year}" for agg in chart_agg if agg['month_group']]
    chart_sales = [agg['total_sale'] for agg in chart_agg if agg['month_group']]
    chart_salary = [agg['total_salary'] for agg in chart_agg if agg['month_group']]
    chart_data = {
        "labels": chart_labels,
        "sales": chart_sales,
        "salaries": chart_salary,
    }

    # --- Chart 2: Expenses by Month (Python aggregation) ---
    monthly_expenses_grouped = defaultdict(float)
    for expense in expenses_qs:
        if expense.expense_date:
            month_label = f"{months[expense.expense_date.month]} {expense.expense_date.year}"
            monthly_expenses_grouped[month_label] += float(expense.amount or 0)
    
    # For the expenses chart we use the union of months from sales and expenses
    all_expense_months_from_expenses = set(monthly_expenses_grouped.keys())
    all_months_in_range = set(chart_labels) | all_expense_months_from_expenses
    expense_chart_labels = sorted(list(all_months_in_range), key=month_year_key)
    expense_chart_data_values = [monthly_expenses_grouped.get(month, 0) for month in expense_chart_labels]
    expense_chart_data = {
        "labels": expense_chart_labels,
        "data": expense_chart_data_values,
    }

    # --- Chart 3 & 4: Manager Sales and Expenses by Type (Python aggregation) ---
    manager_grouped = defaultdict(lambda: defaultdict(float))
    for sale in sales:
        if sale.closing_date and sale.manager:
            month_label = f"{months[sale.closing_date.month]} {sale.closing_date.year}"
            manager_grouped[str(sale.manager)][month_label] += float(sale.sale or 0)

    expense_type_grouped = defaultdict(lambda: defaultdict(float))
    for expense in expenses_qs.select_related('expense_type'):
        if expense.expense_date and expense.expense_type:
            month_label = f"{months[expense.expense_date.month]} {expense.expense_date.year}"
            expense_type_grouped[str(expense.expense_type.name)][month_label] += float(expense.amount or 0)

    # For the "Sales by managers" and "Salary" charts we use only months from sales
    all_sales_months = set(month for manager_data in manager_grouped.values() for month in manager_data)
    # For the "Expense types" chart we use the union of months from sales and expenses
    all_expense_months = set(month for expense_data in expense_type_grouped.values() for month in expense_data)
    all_months = sorted(list(all_sales_months | all_expense_months), key=month_year_key)
    # Separate list of months only for sales and salary charts (without months from expenses)
    all_sales_only_months = sorted(list(all_sales_months), key=month_year_key)

    manager_chart_datasets = []
    for manager, monthly_sales in manager_grouped.items():
        data = [monthly_sales.get(month, 0) for month in all_sales_only_months]
        manager_chart_datasets.append({
            "label": manager,
            "data": data,
        })

    expense_type_datasets = []
    for expense_type_name, monthly_expenses in expense_type_grouped.items():
        data = [monthly_expenses.get(month, 0) for month in all_months]
        expense_type_datasets.append({
            "label": expense_type_name,
            "data": data,
        })

    chart_data_by_manager = {
        "labels": all_sales_only_months,
        "datasets": manager_chart_datasets
    }

    salary_manager_grouped = defaultdict(lambda: defaultdict(float))
    for sale in sales:
        if sale.closing_date and sale.manager:
            month_label = f"{months[sale.closing_date.month]} {sale.closing_date.year}"
            salary_manager_grouped[str(sale.manager)][month_label] += float(sale.salary or 0)

    salary_chart_datasets = []
    for manager, monthly_salary in salary_manager_grouped.items():
        data = [monthly_salary.get(month, 0) for month in all_sales_only_months]
        salary_chart_datasets.append({
            "label": manager,
            "data": data,
        })

    salary_chart_data_by_manager = {
        "labels": all_sales_only_months,
        "datasets": salary_chart_datasets
    }
    
    expense_type_chart_data = {
        "labels": all_months,
        "datasets": expense_type_datasets
    }

    sales_manager_ids = Sale.objects.values_list('manager_id', flat=True).distinct()

    managers = CrmUser.objects.filter(id__in=sales_manager_ids).annotate(
        full_name=Concat(F('last_name'), Value(' '), F('name'), output_field=CharField())
    ).values_list('id', 'full_name').distinct()
    
    # All available years (not from the filtered sales queryset)
    years = Sale.objects.order_by().values_list('closing_date__year', flat=True).distinct()
    years = sorted(filter(None, years))

    context = {
        "sales": sales,
        "managers": managers,
        "months": months,
        "request": request,
        "selected_year": selected_year,
        "is_admin": is_admin,
        "total_sales": total_sales,
        "total_expenses": total_expenses,
        "salary_paid": salary_paid,
        "salary_left": salary_left,
        "years": years,
        "grouped_sales": grouped_sales,
        "chart_data": chart_data,
        "chart_data_by_manager": chart_data_by_manager,
        "salary_chart_data_by_manager": salary_chart_data_by_manager,
        "expense_chart_data": expense_chart_data,
        "expense_type_chart_data": expense_type_chart_data,
        'sort_by': sort_by,
        'order': order,
    }

    return render(request, 'salary/sales_list.html', context)



def export_salary_excel(request):
    manager_id = request.GET.get('manager')
    month = request.GET.get('month')
    year = request.GET.get('year')

    salary_qs = SalaryPayment.objects.select_related('manager').order_by('-payment_datetime')

    # Determine which year is selected for the filter
    current_year = datetime.now().year
    if year == "":
        selected_year = ""  # User selected "All"
    elif not year:
        selected_year = str(current_year)  # Default to current year
    else:
        selected_year = year  # A specific year was selected

    # Apply filters
    if selected_year:
        try:
            y = int(selected_year)
            tz = timezone.get_current_timezone()
            if month:
                m = int(month)
                start, end = get_month_date_range(y, m)
            else:
                start = datetime(y, 1, 1, tzinfo=tz)
                end = datetime(y + 1, 1, 1, tzinfo=tz)
            salary_qs = salary_qs.filter(payment_datetime__gte=start, payment_datetime__lt=end)
        except (ValueError, TypeError):
            # Ignore invalid year/month values
            pass

    if manager_id:
        try:
            salary_qs = salary_qs.filter(manager_id=int(manager_id))
        except (ValueError, TypeError):
            # Ignore invalid manager ID
            pass

    headers = [gettext_lazy("Manager"), gettext_lazy("Payment date"), gettext_lazy("Amount")]
    title = gettext_lazy("Payments")
    filename_prefix = "Salary_payments"

    def salary_row_data_extractor(payment):
        manager_name = f"{payment.manager.last_name} {payment.manager.name}"
        date_str = timezone.localtime(payment.payment_datetime).strftime("%d.%m.%Y %H:%M")
        amount = float(payment.amount)
        return [manager_name, date_str, amount]

    return _generate_excel_response(salary_qs, headers, title, filename_prefix, salary_row_data_extractor)


@login_required
@admin_required
def salary_payment_list(request):
    manager_id = request.GET.get('manager')
    month = request.GET.get('month')
    year = request.GET.get('year')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    filter_type = request.GET.get('filter_type', 'month')

    # --- FIX: Get all managers and years for dropdowns from the complete dataset ---
    # This ensures the filter options are always complete.
    all_payments_for_filters = SalaryPayment.objects.all()
    all_payment_manager_ids = all_payments_for_filters.values_list('manager_id', flat=True).distinct()
    all_managers_for_filter = CrmUser.objects.filter(id__in=all_payment_manager_ids).annotate(
        full_name=Concat(F('last_name'), Value(' '), F('name'), output_field=CharField())
    ).values_list('id', 'full_name')
    # Available years
    years = sorted({dt.year for dt in all_payments_for_filters.values_list('payment_datetime', flat=True) if dt}, reverse=True)
    # --- END FIX ---

    payments = SalaryPayment.objects.select_related('manager').order_by('-payment_datetime')
    
    # Filter by manager
    if manager_id:
        payments = payments.filter(manager_id=manager_id)
    
    # -------------------------------
    # Date Filtering Logic
    # -------------------------------
    current_year = datetime.now().year
    if year == "":
        selected_year = ""          # user selected "All"
    elif not year:
        selected_year = str(current_year)   # first load — default to current year
    else:
        selected_year = year        # a specific year was selected
    
    if filter_type == 'date_range':
        date_start, date_end = parse_date_range(date_from, date_to)
        if date_start and date_end:
            payments = payments.filter(payment_datetime__gte=date_start, payment_datetime__lte=date_end)
    else:  # Default to month filtering
        if selected_year != "":
            try:
                y = int(selected_year)
                if month:
                    try:
                        m = int(month)
                        start, end = get_month_date_range(y, m)
                        payments = payments.filter(payment_datetime__gte=start, payment_datetime__lt=end)
                    except (ValueError, TypeError):
                        pass  # Ignore invalid month formats
                else:
                    tz = timezone.get_current_timezone()
                    start = datetime(y, 1, 1, tzinfo=tz)
                    end = datetime(y + 1, 1, 1, tzinfo=tz)
                    payments = payments.filter(payment_datetime__gte=start, payment_datetime__lt=end)
            except (ValueError, TypeError):
                pass  # Ignore invalid year formats

    total_paid = payments.aggregate(Sum('amount'))['amount__sum'] or 0

    grouped_payments = defaultdict(lambda: {'payments': [], 'total_amount': 0})
    for payment in payments:
        if payment.payment_datetime:
            ym_key = (payment.payment_datetime.year, payment.payment_datetime.month)
            row = cast(dict[str, Any], grouped_payments[ym_key])
            row['payments'].append(payment)
            row['total_amount'] += payment.amount or 0
            
    grouped_payments = dict(sorted(grouped_payments.items()))

    context = {
        'payments': payments,
        'grouped_payments': grouped_payments,
        'managers': all_managers_for_filter,
        'months': get_months(),
        'years': years,
        'request': request,
        'selected_year': selected_year,
        'total_paid': total_paid,
    }
    return render(request, 'salary/salary_payment_list.html', context)

@login_required
@admin_required
def production_expense_list(request):
    employee_id = request.GET.get('employee')
    expense_type_id = request.GET.get('expense_type')
    month = request.GET.get('month')
    year = request.GET.get('year')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    filter_type = request.GET.get('filter_type', 'month')

    # Data for filters
    all_employees = Employee.objects.all().order_by('name')
    all_expense_types = ExpenseType.objects.all().order_by('name')
    years = sorted({dt.year for dt in ProductionExpense.objects.values_list('expense_date', flat=True) if dt}, reverse=True)

    expenses = ProductionExpense.objects.select_related('employee', 'expense_type').order_by('-expense_date')

    # Filter by employee and expense type
    if employee_id:
        expenses = expenses.filter(employee_id=employee_id)
    
    if expense_type_id:
        expenses = expenses.filter(expense_type_id=expense_type_id)

    # -------------------------------
    # Date Filtering Logic
    # -------------------------------
    # By default, no year is selected (which means "All")
    selected_year = year if year is not None else '' 

    # On initial load (no filter params), default to current year.
    is_initial_load = all(p is None for p in [employee_id, expense_type_id, month, year, date_from, date_to])
    if is_initial_load:
        current_year = datetime.now().year
        expenses = expenses.filter(expense_date__year=current_year)
        selected_year = str(current_year)
    elif filter_type == 'date_range':
        date_start, date_end = parse_date_range(date_from, date_to)
        if date_start and date_end:
            expenses = expenses.filter(expense_date__gte=date_start, expense_date__lte=date_end)
        selected_year = '' # Reset year selection in date range mode
    else:  # Month/year filtering
        # Filter by year if a specific year is chosen (not 'All', which is '')
        if selected_year and selected_year.isdigit():
            try:
                expenses = expenses.filter(expense_date__year=int(selected_year))
            except (ValueError, TypeError):
                pass # Ignore invalid year
        
        # Filter by month if a specific month is chosen
        if month and month.isdigit():
            try:
                expenses = expenses.filter(expense_date__month=int(month))
            except (ValueError, TypeError):
                pass # Ignore invalid month

    total_paid = expenses.aggregate(Sum('amount'))['amount__sum'] or 0
    
    grouped_expenses = defaultdict(lambda: {'expenses': [], 'total_amount': 0})
    for expense in expenses:
        if expense.expense_date:
            ym_key = (expense.expense_date.year, expense.expense_date.month)
            row = cast(dict[str, Any], grouped_expenses[ym_key])
            row['expenses'].append(expense)
            row['total_amount'] += expense.amount or 0
            
    grouped_expenses = dict(sorted(grouped_expenses.items()))

    context = {
        'expenses': expenses,
        'grouped_expenses': grouped_expenses,
        'employees': all_employees,
        'expense_types': all_expense_types,
        'months': get_months(),
        'years': years,
        'request': request,
        'selected_year': selected_year,
        'total_paid': total_paid,
    }
    return render(request, 'salary/production_expense_list.html', context)

@login_required
@admin_required
def production_expense_create(request):
    if request.method == 'POST':
        post_data = request.POST.copy()
        if 'amount' in post_data:
            post_data['amount'] = normalize_amount(post_data['amount'])
        form = ProductionExpenseForm(post_data)
        if form.is_valid():
            expense = form.save(commit=False)
            if not expense.expense_date:
                expense.expense_date = timezone.now()
            expense.save()
            if is_ajax(request):
                return JsonResponse({'success': True})
            messages.success(request, gettext_lazy("Expense has been added successfully."))
            return redirect('production_expense_list')
        else:
            if is_ajax(request):
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = ProductionExpenseForm()
    
    if is_ajax(request):
        return render(request, 'salary/production_expense_form_content.html', {
            'form': form,
            'title': gettext_lazy("Add expense"),
        })
    
    return render(request, 'salary/production_expense_form.html', {
        'form': form,
        'title': gettext_lazy("Add expense"),
    })

@login_required
@admin_required
def salary_payment_edit(request, pk):
    payment = get_object_or_404(SalaryPayment, pk=pk)
    remaining_salary = None
    if payment.manager:
        remaining_salary = _get_remaining_salary_for_manager(payment.manager.id)

    if request.method == 'POST':
        post_data = request.POST.copy()
        if 'amount' in post_data:
            post_data['amount'] = normalize_amount(post_data['amount'])
        form = SalaryPaymentForm(post_data, instance=payment)

        if form.is_valid():
            payment = form.save(commit=False)
            # This line incorrectly updates the payment date on every edit.
            # The date should only be changed if the user explicitly changes it in the form.
            # payment.payment_datetime = timezone.now()
            payment.save()
            if is_ajax(request):
                return JsonResponse({'success': True})
            return redirect('salary_payment_list')
        else:
            if is_ajax(request):
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = SalaryPaymentForm(instance=payment)
        # Display amount with spaces (for user convenience)
        if payment.amount is not None:
            form.initial['amount'] = f"{payment.amount:,.2f}".replace(",", " ")
        if payment.payment_datetime:
            form.initial['payment_datetime'] = payment.payment_datetime.strftime('%Y-%m-%dT%H:%M')

    if is_ajax(request):
        return render(request, 'salary/salary_payment_form_content.html', {
            'form': form,
            'payment': payment,
            'action_url': reverse('salary_payment_edit', kwargs={'pk': pk}),
            'title': gettext_lazy("Edit payment #%(pk)s") % {"pk": payment.pk},
            'remaining_salary': remaining_salary
        })

    return render(request, 'salary/salary_payment_form.html', {
        'form': form,
        'payment': payment,
        'title': gettext_lazy("Edit payment #%(pk)s") % {"pk": payment.pk},
        'remaining_salary': remaining_salary
    })

@login_required
@admin_required
def production_expense_edit(request, pk):
    expense = get_object_or_404(ProductionExpense, pk=pk)
    if request.method == 'POST':
        post_data = request.POST.copy()
        if 'amount' in post_data:
            post_data['amount'] = normalize_amount(post_data['amount'])
        form = ProductionExpenseForm(post_data, instance=expense)
        if form.is_valid():
            form.save()
            if is_ajax(request):
                return JsonResponse({'success': True})
            messages.success(request, gettext_lazy("Expense has been updated successfully."))
            return redirect('production_expense_list')
        else:
            if is_ajax(request):
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = ProductionExpenseForm(instance=expense)
        if expense.amount is not None:
            form.initial['amount'] = f"{expense.amount:,.2f}".replace(",", " ")

    if is_ajax(request):
        return render(request, 'salary/production_expense_form_content.html', {
            'form': form,
            'title': gettext_lazy("Edit expense #%(pk)s") % {"pk": expense.pk},
        })

    return render(request, 'salary/production_expense_form.html', {
        'form': form,
        'title': gettext_lazy("Edit expense #%(pk)s") % {"pk": expense.pk},
    })

@login_required
@admin_required
def export_production_excel(request):
    # Get filter parameters from the request
    employee_id = request.GET.get('employee')
    expense_type_id = request.GET.get('expense_type')
    month = request.GET.get('month')
    year = request.GET.get('year')

    # Base queryset
    expenses = ProductionExpense.objects.select_related('employee', 'expense_type').order_by('-expense_date')

    # Filtering logic (copied from production_expense_list)
    current_year = datetime.now().year
    if year == "":
        selected_year = ""
    elif not year:
        selected_year = str(current_year)
    else:
        selected_year = year

    if selected_year:
        try:
            y = int(selected_year)
            tz = timezone.get_current_timezone()
            if month:
                m = int(month)
                start, end = get_month_date_range(y, m)
            else:
                start = datetime(y, 1, 1, tzinfo=tz)
                end = datetime(y + 1, 1, 1, tzinfo=tz)
            expenses = expenses.filter(expense_date__gte=start, expense_date__lt=end)
        except (ValueError, TypeError):
            pass
    elif month:
        try:
            m = int(month)
            expenses = expenses.filter(expense_date__month=m)
        except (ValueError, TypeError):
            pass

    if employee_id:
        expenses = expenses.filter(employee_id=employee_id)
    
    if expense_type_id:
        expenses = expenses.filter(expense_type_id=expense_type_id)

    headers = [
        gettext_lazy("Employee"),
        gettext_lazy("Expense type"),
        gettext_lazy("Date"),
        gettext_lazy("Amount"),
        gettext_lazy("Comment"),
    ]
    title = gettext_lazy("Production expenses")
    filename_prefix = "Production_expenses"
    column_widths = {1: 25, 2: 25, 3: 20, 4: 18, 5: 50}

    def expense_row_data_extractor(expense):
        return [
            expense.employee.name,
            expense.expense_type.name,
            timezone.localtime(expense.expense_date).strftime("%d.%m.%Y %H:%M"),
            float(expense.amount),
            expense.comment
        ]

    return _generate_excel_response(expenses, headers, title, filename_prefix, expense_row_data_extractor, column_widths)

def _generate_excel_response(queryset, headers, title, filename_prefix, row_data_extractor, column_widths=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise ValueError("Workbook has no active sheet")
    title_str = force_str(title) if title is not None else "Sheet1"
    title_str = re.sub(r'[\\*?:/\[\]]', '_', title_str).strip()
    ws.title = (title_str or "Sheet1")[:31]

    header_font = Font(name="Calibri", size=14, bold=True, color="333333")
    thin_border = Border(
        left=Side(style='thin', color="AAAAAA"),
        right=Side(style='thin', color="AAAAAA"),
        top=Side(style='thin', color="AAAAAA"),
        bottom=Side(style='thin', color="AAAAAA")
    )
    header_colors = ["E0E0E0", "D0D0D0", "C0C0C0", "B0B0B0", "A0A0A0"] # Extend as needed

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=force_str(header))
        cell.font = header_font
        fill_color = header_colors[col_num-1] if col_num-1 < len(header_colors) else "C0C0C0"
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, obj in enumerate(queryset, start=2):
        row_data = row_data_extractor(obj)
        fill_color = "EDEDED" if row_idx % 2 == 0 else "FFFFFF"
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_num, value=value)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=14)

            if isinstance(value, (int, float, Decimal)):
                cell.number_format = '# ##0.00'
                cell.alignment = Alignment(horizontal="right")

    if column_widths:
        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width
    else:
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            first_col = column_cells[0].column
            col_letter = get_column_letter(first_col if first_col is not None else 1)
            ws.column_dimensions[col_letter].width = length + 16

    today_file_str = timezone.localdate().strftime("%d-%m-%Y")
    filename = f"{filename_prefix}_{today_file_str}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"

    wb.save(response)
    return response

def _get_remaining_salary_for_manager(manager_id):
    """Helper function to calculate remaining salary for a given manager."""
    try:
        manager = CrmUser.objects.get(id=int(manager_id))
        total_salary = Sale.objects.filter(manager=manager).aggregate(Sum('salary'))['salary__sum'] or 0
        salary_paid = SalaryPayment.objects.filter(manager=manager).aggregate(Sum('amount'))['amount__sum'] or 0
        return total_salary - salary_paid
    except (CrmUser.DoesNotExist, ValueError, TypeError):
        return None

@login_required
@admin_required
def salary_payment_create(request):
    manager_id = request.POST.get('manager') or request.GET.get('manager')
    remaining_salary = None
    
    if manager_id:
        remaining_salary = _get_remaining_salary_for_manager(manager_id)

    if request.method == 'POST':
        post_data = request.POST.copy()
        if 'amount' in post_data:
            post_data['amount'] = normalize_amount(post_data['amount'])
        form = SalaryPaymentForm(post_data)
        if form.is_valid():
            payment = form.save(commit=False)
            # If payment_datetime is empty, set the current time
            if not payment.payment_datetime:
                payment.payment_datetime = timezone.now()
            payment.save()
            if is_ajax(request):
                return JsonResponse({'success': True})
            return redirect('salary_payment_list')
        else:
            if is_ajax(request):
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        initial_data = {'payment_datetime': timezone.now().strftime('%Y-%m-%dT%H:%M')}
        if manager_id:
            initial_data['manager'] = manager_id
        form = SalaryPaymentForm(initial=initial_data)

    if is_ajax(request):
        return render(request, 'salary/salary_payment_form_content.html', {
            'form': form,
            'remaining_salary': remaining_salary,
            'action_url': reverse('salary_payment_create'),
            'title': gettext_lazy("Add payment"),
        })

    return render(request, 'salary/salary_payment_form.html', {
        'form': form,
        'remaining_salary': remaining_salary,
        'title': gettext_lazy("Add payment"),
    })


@require_GET
def get_remaining_salary(request):
    manager_id = request.GET.get('manager_id')
    remaining_salary = _get_remaining_salary_for_manager(manager_id)
    if remaining_salary is not None:
        return JsonResponse({"remaining_salary": f"{remaining_salary:.2f}"})
    else:
        return JsonResponse({"remaining_salary": "0.00"})

@login_required
@admin_required
def employee_create(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            employee = form.save()
            if is_ajax(request):
                return JsonResponse({'success': True, 'id': employee.id, 'name': employee.name})
            messages.success(request, gettext_lazy("Employee '%(name)s' has been added successfully.") % {
                "name": form.cleaned_data['name']
            })
            return redirect('production_expense_list')
        else:
            if is_ajax(request):
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = EmployeeForm()
    
    if is_ajax(request):
        return render(request, 'salary/simple_form_content.html', {
            'form': form,
            'title': gettext_lazy("Add employee"),
            'action_url': reverse('employee_create'),
        })
    
    return render(request, 'salary/simple_form.html', {
        'form': form,
        'title': gettext_lazy("Add employee"),
    })

@login_required
@admin_required
def expense_type_create(request):
    if request.method == 'POST':
        form = ExpenseTypeForm(request.POST)
        if form.is_valid():
            expense_type = form.save()
            if is_ajax(request):
                return JsonResponse({'success': True, 'id': expense_type.id, 'name': expense_type.name})
            messages.success(request, gettext_lazy("Expense type '%(name)s' has been added successfully.") % {
                "name": form.cleaned_data['name']
            })
            return redirect('production_expense_list')
        else:
            if is_ajax(request):
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        form = ExpenseTypeForm()

    if is_ajax(request):
        return render(request, 'salary/simple_form_content.html', {
            'form': form,
            'title': gettext_lazy("Add expense type"),
            'action_url': reverse('expense_type_create'),
        })

    return render(request, 'salary/simple_form.html', {
        'form': form,
        'title': gettext_lazy("Add expense type"),
    })





@login_required
@admin_required  
def register(request):
    manager_id = request.GET.get('manager_id')

    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            if is_ajax(request):
                return JsonResponse({'success': True})
            messages.success(request, gettext_lazy("User has been registered successfully."))
            return redirect("users_list")
        else:
            if is_ajax(request):
                return JsonResponse({'success': False, 'errors': form.errors})
    else:
        # If manager_id is provided, pre-fill it into the form's initial data
        initial_data = {}
        if manager_id:
            initial_data['manager'] = manager_id
        form = RegistrationForm(initial=initial_data)

    if is_ajax(request):
        return render(request, "salary/register_form_content.html", {"form": form, 'action_url': reverse('register')})

    return render(request, "salary/register.html", {"form": form})


@login_required
@admin_required
def users_list(request):
    users = CrmUser.objects.select_related('django_user').all()
    return render(request, 'salary/users_list.html', {'users': users})


@login_required
@admin_required
def delete_user_account(request, manager_id):
    manager = get_object_or_404(CrmUser, id=manager_id)
    if manager.django_user:
        # Delete the related Django user (login and password)
        # The on_delete=models.SET_NULL on the User.django_user field
        # automatically handles setting the field to None.
        manager.django_user.delete()
        
        if manager.is_admin:
            CrmUser.objects.filter(pk=manager.pk).update(is_admin=False)

        messages.success(request, gettext_lazy("Login and password have been removed for user %(last)s %(first)s.") % {
            "last": manager.last_name,
            "first": manager.name,
        })
    else:
        messages.warning(request, gettext_lazy("This manager does not have a registered user."))
    return redirect('users_list')

def register_with_manager(request, manager_id):
    """
    Redirect to the registration page, passing the manager id.
    On the registration page this id can be used to
    link the created django_user to the selected manager.
    """
    return redirect(f"/register?manager_id={manager_id}")


# AI Analysis views — DISABLED: local neural-network hooks commented out
# ai_analysis_view = ai_views.ai_analysis_view
# ai_analyze_data = ai_views.ai_analyze_data
# ai_generate_chart = ai_views.ai_generate_chart
# ai_check_model_status = ai_views.ai_check_model_status
